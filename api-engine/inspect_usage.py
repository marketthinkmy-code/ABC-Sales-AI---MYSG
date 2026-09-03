"""
只讀探勘『軟體使用數據』表的結構(給後續 買家×使用率 交叉用)。
只印:欄位名 · 列數 · 各『低基數欄位』的值分佈(行業/方案/狀態這種)。
高基數欄位(email/姓名/id)只印非空筆數,不印任何值 —— 不外洩個資。

env: GOOGLE_SA_JSON, USAGE_SHEET_ID, USAGE_GID(選填)
"""
import os, io, csv, json
from collections import Counter
from google.oauth2 import service_account
from googleapiclient.discovery import build

USAGE_ID = os.environ["USAGE_SHEET_ID"]
GID = os.environ.get("USAGE_GID") or ""
SCOPES = ["https://www.googleapis.com/auth/drive.readonly",
          "https://www.googleapis.com/auth/spreadsheets.readonly"]


def _sci(v):
    return "" if v is None else str(v).strip()


def load():
    info = json.loads(os.environ["GOOGLE_SA_JSON"])
    print("SA client_email(要把表分享給這個帳號才讀得到):", info.get("client_email"))
    creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    drive = build("drive", "v3", credentials=creds)
    # 這份是「上傳的 .xlsx」,不是原生 Google Sheet → 直接下載二進位用 openpyxl 讀。
    import openpyxl
    data = drive.files().get_media(fileId=USAGE_ID).execute()
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    print(f"分頁清單:{wb.sheetnames}")
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [[_sci(c) for c in r] for r in ws.iter_rows(values_only=True)]
        out[name] = rows
    return out


def inspect(name, vals):
    if not vals:
        print(f"\n===== 分頁「{name}」:空 =====")
        return
    hdr_i = max(range(min(6, len(vals))), key=lambda i: sum(1 for c in vals[i] if c.strip()))
    hdr = vals[hdr_i]
    rows = vals[hdr_i + 1:]
    print(f"\n===== 分頁「{name}」· 共 {len(vals)} 列 · 表頭第 {hdr_i+1} 列 · 資料 {len(rows)} 筆 =====")
    for j, nm in enumerate(hdr):
        if not nm.strip():
            continue
        col = [r[j].strip() for r in rows if j < len(r) and r[j].strip()]
        distinct = len(set(col))
        low = 0 < distinct <= 25
        note = "  ← 分類欄" if low else ("  (高基數/略過值)" if distinct else "  (空)")
        print(f"  [{j:>2}] {nm[:28]:<28} 非空 {len(col):>4} · 相異 {distinct:>4}{note}")
        if low:
            for v, c in Counter(col).most_common(25):
                print(f"          {c:>4} × {v[:44]}")


def dump(name, vals, max_rows=60, max_cols=26):
    print(f"\n===== 分頁「{name}」原始格 · {len(vals)} 列 =====")
    for i, row in enumerate(vals[:max_rows]):
        cells = [c for c in row[:max_cols]]
        if not any(c.strip() for c in cells):
            continue
        line = " | ".join((c[:16] if c else "·") for c in cells).rstrip(" |·").rstrip()
        if line.strip():
            print(f"  r{i:>2}: {line}")


def run():
    sheets = load()
    mode = os.environ.get("MODE") or "dump"
    for name, vals in sheets.items():
        if mode == "inspect":
            inspect(name, vals)
        else:
            dump(name, vals)


if __name__ == "__main__":
    run()
