#!/usr/bin/env python3
"""Split the Salesbot secured-sales export into MY / SG / TW buckets.

Only rows whose "Lead Source" mentions MTC are kept — those are the ones we
need to record. Region is decided from the phone number:

    +886 / 886...        -> TW
    +60  / 60... / 0xx   -> MY   (local Malaysian format has no country code)
    +65  / 65...         -> SG
    anything else        -> Others (flagged for manual confirmation)

Usage:
    python scripts/split_mtc_sales_by_region.py <input.csv> [output.xlsx]
"""

import csv
import re
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_OUT = 'data/MTC_Secured_Sales_by_Region.xlsx'

# Rows that need a human to decide the region, keyed by the Name column.
NOTES = {
    'GRAND VISION X SDN BHD': 'No phone number in source — company name suggests MY, needs manual confirmation',
    'Adrian': '1st number +66 (Thailand), 2nd number +60 (Malaysia) — needs manual confirmation',
    'RYION': '+852 (Hong Kong)',
}

FONT = 'Arial'


def first_num(phone):
    """A few rows list two contacts; the first one decides the region."""
    parts = [x for x in re.split(r'//|/|,', phone) if re.search(r'\d', x)]
    return parts[0].strip() if parts else ''


def norm(phone):
    return re.sub(r'\D', '', first_num(phone))


def classify(phone):
    d = norm(phone)
    if not d:
        return 'Others'
    if d.startswith('886'):
        return 'TW'
    if d.startswith('60'):
        return 'MY'
    if d.startswith('65'):
        return 'SG'
    if d.startswith('0'):
        return 'MY'
    return 'Others'


def build(src, out):
    rows = list(csv.reader(open(src, newline='', encoding='utf-8-sig')))
    hdr = rows[0][:13]
    data = [r for r in rows[1:] if len(r) >= 13 and any(x.strip() for x in r[:13])]
    mtc = [r[:13] for r in data if 'mtc' in r[5].lower()]

    buckets = {'MY': [], 'SG': [], 'TW': [], 'Others': []}
    for r in mtc:
        buckets[classify(r[3])].append(r)

    hdr_font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    cell_font = Font(name=FONT, size=10)
    note_font = Font(name=FONT, size=10, italic=True, color='C00000')
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    sm = wb.active
    sm.title = 'Summary'
    out_hdr = hdr + ['Region', 'Phone (normalised)', 'Note']

    def add_sheet(name, recs):
        ws = wb.create_sheet(name)
        ws.append(out_hdr)
        for c in ws[1]:
            c.font, c.fill, c.border = hdr_font, hdr_fill, border
            c.alignment = Alignment(horizontal='center', vertical='center')
        for r in recs:
            ws.append(r + [name, norm(r[3]), NOTES.get(r[1].strip(), '')])
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font, c.border = cell_font, border
            if row[15].value:
                row[15].font = note_font
        if not recs:
            ws.cell(row=1, column=1).comment = Comment(
                'No records: no Taiwan (+886) phone numbers were found among the MTC rows '
                'in the source file. Header kept so future TW sales can be appended here.',
                'Claude')
        for i, w in enumerate([6, 32, 32, 26, 20, 20, 18, 16, 16, 14, 12, 14, 26, 10, 18, 60], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        if recs:
            ws.auto_filter.ref = f'A1:{get_column_letter(len(out_hdr))}{len(recs) + 1}'

    for name in ('MY', 'SG', 'TW', 'Others'):
        add_sheet(name, buckets[name])

    sm['A1'] = 'MTC Secured Sales — split by region (phone number)'
    sm['A1'].font = Font(name=FONT, bold=True, size=14)
    meta = [('Source file', src),
            ('Filter applied', 'Lead Source contains "MTC" (case-insensitive)'),
            ('Total rows in source', len(data)),
            ('Rows kept (MTC)', len(mtc))]
    for i, (k, v) in enumerate(meta):
        sm.cell(row=3 + i, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        sm.cell(row=3 + i, column=2, value=v).font = cell_font

    for col, title in ((1, 'Region'), (2, 'Records'), (3, 'Phone rule')):
        c = sm.cell(row=8, column=col, value=title)
        c.font, c.fill, c.border = hdr_font, hdr_fill, border
        c.alignment = Alignment(horizontal='center')

    rules = [('MY', 'Starts with 60 / +60, or local 0xx format'),
             ('SG', 'Starts with 65 / +65'),
             ('TW', 'Starts with 886 / +886'),
             ('Others', 'Any other country code (e.g. +852, +66) or no phone number')]
    row = 9
    for label, rule in rules:
        sm.cell(row=row, column=1, value=label).font = cell_font
        sm.cell(row=row, column=2, value=f'=COUNTA({label}!A2:A2000)').font = cell_font
        sm.cell(row=row, column=3, value=rule).font = cell_font
        for col in (1, 2, 3):
            sm.cell(row=row, column=col).border = border
        row += 1
    sm.cell(row=row, column=1, value='TOTAL').font = Font(name=FONT, bold=True, size=10)
    sm.cell(row=row, column=2, value=f'=SUM(B9:B{row - 1})').font = Font(name=FONT, bold=True, size=10)
    for col in (1, 2, 3):
        sm.cell(row=row, column=col).border = border

    sm.cell(row=16, column=1, value='Notes').font = Font(name=FONT, bold=True, size=11)
    notes = [
        '1. Region is determined by the FIRST phone number in the "Phone Number" column when a row lists more than one.',
        '2. Local Malaysian formats (e.g. 012-345 6789) have no country code and are treated as MY.',
        '3. No Taiwan (+886) numbers exist in this file, so the TW sheet is empty (header kept for future use).',
        '4. Rows needing manual confirmation are flagged in red in the "Note" column of the Others sheet.',
        '5. "Phone (normalised)" is the first phone number with all non-digit characters removed.',
    ]
    for i, n in enumerate(notes):
        sm.cell(row=17 + i, column=1, value=n).font = cell_font

    sm.column_dimensions['A'].width = 110
    sm.column_dimensions['B'].width = 14
    sm.column_dimensions['C'].width = 58

    wb.save(out)
    return {k: len(v) for k, v in buckets.items()}


if __name__ == '__main__':
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    counts = build(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT)
    print(counts)
